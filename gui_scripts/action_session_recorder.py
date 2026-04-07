"""
从子进程启动起记录 pyautogui 的输入类操作：优先截取「游戏窗口识别」配置的区域（与 capture_game_region 一致），
按该区域原始像素分辨率保存 PNG，并在图上标注屏幕坐标换算后的点击/拖放；写入 events.jsonl。
未配置游戏区域或截取失败时回退为全屏截图（同样不缩放）。

控制中心在停止循环（含 F12）约 1 秒后调用 finalize_session_dir 生成带缩略图的 Excel。

环境变量：QQFARM_ACTION_SESSION_DIR = 本会话根目录（由控制中心创建）。

保留策略（由控制中心在创建新会话前执行）：最多保留 ACTION_SESSION_LOG_KEEP_COUNT 个会话目录。
本会话目录总大小（截图 + jsonl 等）达到 ACTION_SESSION_MAX_TOTAL_BYTES 后不再记录新事件。
"""
from __future__ import annotations

import json
import math
import os
import threading
import time
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

# 与 control_center_core._prune_action_session_logs 的 keep 数量保持一致
ACTION_SESSION_LOG_KEEP_COUNT = 5
ACTION_SESSION_MAX_TOTAL_BYTES = 100 * 1024 * 1024

_session: Optional[Dict[str, Any]] = None
_orig: Dict[str, Callable[..., Any]] = {}


def _root() -> Path:
    assert _session is not None
    return _session["root"]


def _t_rel() -> float:
    assert _session is not None
    return round(time.perf_counter() - float(_session["t0"]), 3)


def _dir_total_bytes(root: Path) -> int:
    total = 0
    try:
        for p in root.rglob("*"):
            if p.is_file():
                try:
                    total += int(p.stat().st_size)
                except OSError:
                    pass
    except OSError:
        pass
    return total


def _clamp_local(lx: int, ly: int, w: int, h: int) -> Tuple[int, int]:
    return max(0, min(w - 1, lx)), max(0, min(h - 1, ly))


def _markers_screen_to_region(
    markers: List[Tuple[int, int]],
    region: Dict[str, int],
) -> List[Tuple[int, int]]:
    rx, ry = int(region["x"]), int(region["y"])
    w, h = int(region["w"]), int(region["h"])
    out: List[Tuple[int, int]] = []
    for sx, sy in markers:
        out.append(_clamp_local(int(sx) - rx, int(sy) - ry, w, h))
    return out


def _line_screen_to_region(
    line: Optional[Tuple[Tuple[int, int], Tuple[int, int]]],
    region: Dict[str, int],
) -> Optional[Tuple[Tuple[int, int], Tuple[int, int]]]:
    if line is None:
        return None
    rx, ry = int(region["x"]), int(region["y"])
    w, h = int(region["w"]), int(region["h"])
    (ax, ay), (bx, by) = line
    return (
        _clamp_local(int(ax) - rx, int(ay) - ry, w, h),
        _clamp_local(int(bx) - rx, int(by) - ry, w, h),
    )


def _capture_frame_for_record(
    markers_screen: List[Tuple[int, int]],
    line_screen: Optional[Tuple[Tuple[int, int], Tuple[int, int]]],
):
    """
    返回 (PIL.Image RGB, markers 与 line 已为与图像一致的坐标系)。
    游戏区域模式：图像为区域原始分辨率，坐标为区域内像素。
    """
    from PIL import Image

    try:
        from game_region_locator import capture_game_region, load_config_region

        region = load_config_region()
        if region and int(region.get("w", 0)) > 0 and int(region.get("h", 0)) > 0:
            import cv2
            import numpy as np

            bgr = capture_game_region()
            if bgr is not None and getattr(bgr, "size", 0) > 0 and bgr.shape[0] > 0 and bgr.shape[1] > 0:
                rgb = cv2.cvtColor(np.asarray(bgr), cv2.COLOR_BGR2RGB)
                im = Image.fromarray(rgb)
                lm = _markers_screen_to_region(markers_screen, region)
                ll = _line_screen_to_region(line_screen, region)
                return im, lm, ll
    except Exception:
        pass

    import pyautogui

    try:
        shot = pyautogui.screenshot()
    except Exception:
        return None, [], None
    if shot.mode != "RGB":
        shot = shot.convert("RGB")
    return shot, list(markers_screen), line_screen


def _draw_drag_arrow(
    draw,
    ax: int,
    ay: int,
    bx: int,
    by: int,
    *,
    line_rgba: Tuple[int, int, int, int] = (255, 100, 0, 255),
    head_len: float = 26.0,
    head_half_w: float = 14.0,
) -> None:
    """在 (bx,by) 端画指向该点的箭头（沿 a→b 方向）。"""
    dx = float(bx - ax)
    dy = float(by - ay)
    ln = math.hypot(dx, dy)
    if ln < 1e-6:
        return
    ux, uy = dx / ln, dy / ln
    px, py = -uy, ux
    bx_f, by_f = float(bx), float(by)
    base_x = bx_f - head_len * ux
    base_y = by_f - head_len * uy
    p1 = (bx, by)
    p2 = (int(base_x + head_half_w * px), int(base_y + head_half_w * py))
    p3 = (int(base_x - head_half_w * px), int(base_y - head_half_w * py))
    draw.line([(ax, ay), (int(base_x), int(base_y))], fill=line_rgba, width=8)
    draw.polygon([p1, p2, p3], fill=line_rgba, outline=line_rgba)


def _annotate_shot(
    base,
    markers: List[Tuple[int, int]],
    line: Optional[Tuple[Tuple[int, int], Tuple[int, int]]] = None,
    *,
    drag_track_style: bool = False,
):
    from PIL import ImageDraw

    draw = ImageDraw.Draw(base, "RGBA")
    if drag_track_style and line:
        a, b = line
        ax, ay = int(a[0]), int(a[1])
        bx, by = int(b[0]), int(b[1])
        _draw_drag_arrow(draw, ax, ay, bx, by)
        rs, re = 11, 11
        draw.ellipse((ax - rs, ay - rs, ax + rs, ay + rs), outline=(0, 180, 80, 255), width=4)
        draw.ellipse((bx - re, by - re, bx + re, by + re), outline=(255, 40, 40, 255), width=4)
        ang = math.degrees(math.atan2(by - ay, bx - ax))
        from PIL import ImageFont

        try:
            font = ImageFont.load_default()
        except Exception:
            font = None
        label = f"→ {ang:.0f}°"
        if font:
            draw.text((bx + 12, by - 8), label, fill=(255, 255, 0, 255), font=font)
        else:
            draw.text((bx + 12, by - 8), label, fill=(255, 255, 0, 255))
        return base

    for x, y in markers:
        r = 16
        draw.ellipse((x - r, y - r, x + r, y + r), outline=(255, 0, 0, 255), width=4)
        draw.line((x - 22, y, x + 22, y), fill=(255, 0, 0, 255), width=2)
        draw.line((x, y - 22, x, y + 22), fill=(255, 0, 0, 255), width=2)
    if line:
        a, b = line
        draw.line([a, b], fill=(255, 140, 0, 255), width=5)
        for p, col in ((a, (0, 200, 0, 255)), (b, (255, 0, 0, 255))):
            r = 10
            draw.ellipse((p[0] - r, p[1] - r, p[0] + r, p[1] + r), outline=col, width=3)
    return base


def ensure_action_session_from_env() -> bool:
    """环境变量已设会话目录但尚未 install 时补装（避免叠加图等 append 时 _session 仍为 None）。"""
    d = os.environ.get("QQFARM_ACTION_SESSION_DIR", "").strip()
    if not d:
        return False
    if _session is not None:
        return True
    install_from_env()
    return _session is not None


def append_session_png_bytes(data: bytes, action: str, detail: str) -> None:
    """写入已编码的 PNG 字节并追加 events.jsonl（与 _take_and_save 同一套序号与容量策略）。"""
    if _session is None:
        ensure_action_session_from_env()
    if _session is None or not data:
        return
    sz = len(data)
    with _session["lock"]:
        if bool(_session.get("recording_stopped_size", False)):
            return
        cur_b = int(_session.get("bytes_written", 0))
        if cur_b >= ACTION_SESSION_MAX_TOTAL_BYTES:
            _session["recording_stopped_size"] = True
            if not _session.get("size_cap_logged"):
                _session["size_cap_logged"] = True
                print(
                    f"[action_session] 本会话目录已达 {ACTION_SESSION_MAX_TOTAL_BYTES // (1024 * 1024)}MB 上限，已停止截图与事件记录。",
                    flush=True,
                )
            return
        if cur_b + sz > ACTION_SESSION_MAX_TOTAL_BYTES:
            _session["recording_stopped_size"] = True
            if not _session.get("size_cap_logged"):
                _session["size_cap_logged"] = True
                print(
                    f"[action_session] 再写入将超过 {ACTION_SESSION_MAX_TOTAL_BYTES // (1024 * 1024)}MB 上限，已停止记录。",
                    flush=True,
                )
            return
        _session["seq"] = int(_session["seq"]) + 1
        seq = int(_session["seq"])
        fname = f"{seq:04d}_{action}.png"
        rel = f"shots/{fname}"
        outp = _root() / "shots" / fname

    ev = {
        "seq": seq,
        "t_sec": _t_rel(),
        "action": action,
        "detail": detail,
        "file": rel,
    }
    jline = (json.dumps(ev, ensure_ascii=False) + "\n").encode("utf-8")
    if cur_b + sz + len(jline) > ACTION_SESSION_MAX_TOTAL_BYTES:
        with _session["lock"]:
            _session["seq"] = int(_session["seq"]) - 1
            _session["recording_stopped_size"] = True
            if not _session.get("size_cap_logged"):
                _session["size_cap_logged"] = True
                print(
                    f"[action_session] 再写入将超过 {ACTION_SESSION_MAX_TOTAL_BYTES // (1024 * 1024)}MB 上限，已停止记录。",
                    flush=True,
                )
        return

    try:
        outp.write_bytes(data)
    except Exception:
        with _session["lock"]:
            _session["seq"] = int(_session["seq"]) - 1
        return

    with _session["lock"]:
        try:
            with open(_root() / "events.jsonl", "ab") as f:
                f.write(jline)
                f.flush()
        except Exception:
            try:
                outp.unlink(missing_ok=True)
            except OSError:
                pass
            _session["seq"] = int(_session["seq"]) - 1
            return
        _session["bytes_written"] = int(_session.get("bytes_written", 0)) + sz + len(jline)


def append_session_bgr_image(bgr, action: str, detail: str) -> bool:
    """会话调试：写入一张 BGR 图（不经 pyautogui），例如叠加可视化调试图。成功返回 True。"""
    ensure_action_session_from_env()
    try:
        import numpy as np
        from io import BytesIO

        from PIL import Image

        arr = np.asarray(bgr)
        if arr.ndim != 3 or arr.shape[2] != 3:
            print(
                f"[action_session] append_session_bgr_image 跳过: 形状 {getattr(arr, 'shape', None)} 非 HWx3",
                flush=True,
            )
            return False
        if arr.dtype != np.uint8:
            arr = np.clip(arr, 0, 255).astype(np.uint8)
        rgb = np.ascontiguousarray(arr[:, :, ::-1])
        buf = BytesIO()
        Image.fromarray(rgb, mode="RGB").save(buf, format="PNG", optimize=True)
        data = buf.getvalue()
        if _session is None:
            return False
        seq0 = int(_session["seq"])
        append_session_png_bytes(data, action, detail)
        return int(_session["seq"]) > seq0
    except Exception as exc:
        print(f"[action_session] append_session_bgr_image 失败: {type(exc).__name__}: {exc}", flush=True)
        return False


def _take_and_save(
    action: str,
    detail: str,
    markers: List[Tuple[int, int]],
    line: Optional[Tuple[Tuple[int, int], Tuple[int, int]]] = None,
    *,
    drag_track_style: bool = False,
) -> None:
    if _session is None:
        return
    with _session["lock"]:
        if bool(_session.get("recording_stopped_size", False)):
            return
        cur_b = int(_session.get("bytes_written", 0))
        if cur_b >= ACTION_SESSION_MAX_TOTAL_BYTES:
            _session["recording_stopped_size"] = True
            if not _session.get("size_cap_logged"):
                _session["size_cap_logged"] = True
                print(
                    f"[action_session] 本会话目录已达 {ACTION_SESSION_MAX_TOTAL_BYTES // (1024 * 1024)}MB 上限，已停止截图与事件记录。",
                    flush=True,
                )
            return

    from io import BytesIO

    shot, markers_draw, line_draw = _capture_frame_for_record(markers, line)
    if shot is None:
        return
    _annotate_shot(
        shot,
        [] if drag_track_style and line_draw else markers_draw,
        line_draw,
        drag_track_style=drag_track_style,
    )

    buf = BytesIO()
    try:
        shot.save(buf, format="PNG", optimize=True)
    except Exception:
        return
    data = buf.getvalue()
    append_session_png_bytes(data, action, detail)


def _detail_from_call(name: str, args: tuple, kwargs: dict) -> str:
    kw = {k: v for k, v in kwargs.items() if k in ("button", "clicks", "interval", "duration", "pause")}
    if name in ("click", "doubleClick", "rightClick", "middleClick"):
        x = args[0] if len(args) > 0 else kwargs.get("x")
        y = args[1] if len(args) > 1 else kwargs.get("y")
        extra = f", {kw}" if kw else ""
        return f"({x}, {y}){extra}"
    if name == "moveTo":
        x = args[0] if len(args) > 0 else kwargs.get("x")
        y = args[1] if len(args) > 1 else kwargs.get("y")
        extra = f", {kw}" if kw else ""
        return f"({x}, {y}){extra}"
    if name == "moveRel":
        return f"args={args!r} {kw}"
    if name in ("dragTo", "drag", "dragRel"):
        return f"args={args!r} {kw}"
    if name in ("mouseDown", "mouseUp"):
        return f"{kw}"
    if name == "scroll":
        return f"clicks={args[0] if args else kwargs.get('clicks')} {kw}"
    if name in ("keyDown", "keyUp", "press"):
        return f"{args!r} {kw}"
    if name == "hotkey":
        return f"{args!r}"
    if name in ("typewrite", "write"):
        s = args[0] if args else ""
        if isinstance(s, str) and len(s) > 80:
            s = s[:77] + "..."
        return f"{s!r} {kw}"
    return f"args={args!r} kwargs={kwargs!r}"


def _move_to_duration_seconds(args: tuple, kwargs: dict) -> float:
    d = kwargs.get("duration")
    if d is None and len(args) >= 3:
        d = args[2]
    if d is None:
        return 0.0
    try:
        return float(d)
    except (TypeError, ValueError):
        return 0.0


def _resolve_mouse_pos(args: tuple, kwargs: dict, pa) -> Tuple[int, int]:
    """当前将要操作的屏幕坐标（与 pyautogui 一致：未传 x,y 则用当前光标位置）。"""
    x = kwargs.get("x")
    y = kwargs.get("y")
    if x is None and len(args) >= 1 and args[0] is not None:
        x = args[0]
    if y is None and len(args) >= 2 and args[1] is not None:
        y = args[1]
    if x is not None and y is not None:
        return int(x), int(y)
    pos = pa.position()
    return int(pos[0]), int(pos[1])


def _markers_for_call(name: str, args: tuple, kwargs: dict) -> List[Tuple[int, int]]:
    import pyautogui as pa

    if name in ("click", "doubleClick", "rightClick", "middleClick"):
        x = args[0] if len(args) > 0 else kwargs.get("x")
        y = args[1] if len(args) > 1 else kwargs.get("y")
        if x is not None and y is not None:
            return [(int(x), int(y))]
    if name == "moveTo":
        x = args[0] if len(args) > 0 else kwargs.get("x")
        y = args[1] if len(args) > 1 else kwargs.get("y")
        if x is not None and y is not None:
            return [(int(x), int(y))]
    if name in ("dragTo", "drag"):
        x = args[0] if len(args) > 0 else kwargs.get("x")
        y = args[1] if len(args) > 1 else kwargs.get("y")
        if x is not None and y is not None:
            cur = pa.position()
            return [(int(cur[0]), int(cur[1])), (int(x), int(y))]
    if name == "mouseUp":
        cur = pa.position()
        return [(int(cur[0]), int(cur[1]))]
    return []


def _make_wrapper(name: str, orig: Callable[..., Any]) -> Callable[..., Any]:
    def wrapped(*args: Any, **kwargs: Any):
        if _session is None:
            return orig(*args, **kwargs)
        import pyautogui as pa

        # 先截图再执行操作，避免界面已切换导致「截图滞后」（尤其点击后进拜访/好友等场景）。
        if name == "mouseDown":
            px, py = _resolve_mouse_pos(args, kwargs, pa)
            _session["mouse_down"] = (px, py, str(kwargs.get("button", "left")))
            # 拖放只在 mouseUp 生成一张 drag_track 轨迹图，此处不截图。
            return orig(*args, **kwargs)

        if name == "mouseUp":
            x, y = pa.position()
            down = _session.get("mouse_down")
            line = None
            markers: List[Tuple[int, int]] = [(int(x), int(y))]
            if down and len(down) >= 2:
                ax, ay = int(down[0]), int(down[1])
                line = ((ax, ay), (int(x), int(y)))
                markers = [(ax, ay), (int(x), int(y))]
            dist = 0.0
            if line:
                ax, ay = int(line[0][0]), int(line[0][1])
                dist = math.hypot(float(x - ax), float(y - ay))
            detail_base = _detail_from_call(name, args, kwargs)
            # 松开前截图：左键仍按下。mousedown 后只要起点≠终点就记为 drag_track（小幅拖移不再落一条易被当成「点击」的 mouseUp）。
            if line and dist > 0:
                detail = (
                    f"拖放轨迹 屏幕像素: 起点{line[0]} → 终点{line[1]}，位移约{dist:.1f}px；{detail_base}"
                )
                _take_and_save(
                    "drag_track",
                    detail,
                    [],
                    line,
                    drag_track_style=True,
                )
            else:
                detail = detail_base + (f" 松手位置: ({x},{y})" if not line else f" 拖放: {line[0]} -> {line[1]}")
                _take_and_save("mouseUp", detail, markers, line, drag_track_style=False)
            res = orig(*args, **kwargs)
            _session["mouse_down"] = None
            return res

        # 按住鼠标拖移时不再逐帧记录 moveTo/moveRel（避免拖移过程产生大量中间帧）；duration=0 的 moveTo 也不记。
        if name == "moveTo":
            if _session.get("mouse_down"):
                return orig(*args, **kwargs)
            if _move_to_duration_seconds(args, kwargs) <= 0.0:
                return orig(*args, **kwargs)
        if name == "moveRel" and _session.get("mouse_down"):
            return orig(*args, **kwargs)
        if name == "dragRel" and _session.get("mouse_down"):
            return orig(*args, **kwargs)

        detail = _detail_from_call(name, args, kwargs)
        markers = _markers_for_call(name, args, kwargs)
        line = None
        if name in ("dragTo", "drag") and len(markers) >= 2:
            line = (markers[0], markers[1])
        _take_and_save(name, detail, markers, line)
        return orig(*args, **kwargs)

    return wrapped


def install(session_dir: str) -> None:
    """在导入 game_region_locator 之前调用；会 patch pyautogui。"""
    global _session
    if _session is not None:
        return
    root = Path(session_dir)
    root.mkdir(parents=True, exist_ok=True)
    (root / "shots").mkdir(exist_ok=True)
    _session = {
        "root": root,
        "t0": time.perf_counter(),
        "seq": 0,
        "lock": threading.Lock(),
        "mouse_down": None,
        "bytes_written": _dir_total_bytes(root),
        "recording_stopped_size": False,
        "size_cap_logged": False,
    }
    import pyautogui as pa

    names = [
        "click",
        "doubleClick",
        "rightClick",
        "middleClick",
        "moveTo",
        "moveRel",
        "drag",
        "dragTo",
        "dragRel",
        "mouseDown",
        "mouseUp",
        "scroll",
        "hscroll",
        "keyDown",
        "keyUp",
        "press",
        "hotkey",
        "typewrite",
        "write",
    ]
    for n in names:
        if hasattr(pa, n):
            fn = getattr(pa, n)
            if callable(fn) and n not in _orig:
                _orig[n] = fn
                setattr(pa, n, _make_wrapper(n, fn))
    _take_and_save("session", "任务操作记录已开始（后续每条为自动化输入）", [])


def finalize_session_dir(root: Path) -> Optional[Path]:
    """根据 events.jsonl 生成 timeline.xlsx；若缺少 openpyxl 则跳过。"""
    root = Path(root)
    jsonl = root / "events.jsonl"
    if not jsonl.is_file():
        return None
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Alignment
    except Exception:
        return None

    events: List[dict] = []
    with open(jsonl, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                events.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    if not events:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "操作时间线"
    headers = ("序号", "自开始(秒)", "操作类型", "详情", "截图文件")
    ws.append(headers)
    for c in range(1, 6):
        ws.cell(row=1, column=c).alignment = Alignment(vertical="center", wrap_text=True)

    img_col = 6
    ws.cell(row=1, column=img_col, value="预览图")

    for i, ev in enumerate(events, start=2):
        ws.cell(row=i, column=1, value=ev.get("seq"))
        ws.cell(row=i, column=2, value=ev.get("t_sec"))
        ws.cell(row=i, column=3, value=str(ev.get("action", "")))
        ws.cell(row=i, column=4, value=str(ev.get("detail", "")))
        ws.cell(row=i, column=5, value=str(ev.get("file", "")))
        for c in range(1, 6):
            ws.cell(row=i, column=c).alignment = Alignment(vertical="top", wrap_text=True)
        rel = ev.get("file")
        if rel:
            p = root / str(rel).replace("\\", "/")
            if p.is_file():
                try:
                    xl_img = XLImage(str(p))
                    max_w, max_h = 480, 300
                    if (xl_img.width or 0) > max_w:
                        sc = max_w / float(xl_img.width)
                        xl_img.width = max_w
                        xl_img.height = int((xl_img.height or 1) * sc)
                    if (xl_img.height or 0) > max_h:
                        sc = max_h / float(xl_img.height)
                        xl_img.height = max_h
                        xl_img.width = int((xl_img.width or 1) * sc)
                    anchor = f"{chr(ord('A') + img_col - 1)}{i}"
                    ws.add_image(xl_img, anchor)
                    ws.row_dimensions[i].height = max(90, min(240, (xl_img.height or 100) * 0.78))
                except Exception:
                    pass

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 56
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 70

    out = root / "timeline.xlsx"
    wb.save(str(out))
    return out


def finalize_if_active() -> None:
    if _session is None:
        return
    try:
        finalize_session_dir(_root())
    except Exception:
        pass


def install_from_env() -> None:
    d = os.environ.get("QQFARM_ACTION_SESSION_DIR", "").strip()
    if d:
        install(d)
